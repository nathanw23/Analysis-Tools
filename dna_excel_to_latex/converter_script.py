import pandas as pd
import openpyxl

base_in = r"C:\Users\s1893121\OneDrive - University of Edinburgh\PhD\Laboratory\Experiments\ToeholdInDel_Investigation\InDel_Version3\InDelv3_Strands.xlsx"
base_out = r"C:\Users\s1893121\Downloads"

# Load Excel workbook using openpyxl (to keep track of rich text)
workbook = openpyxl.load_workbook(base_in, rich_text=True)

sheets = workbook.sheetnames

for sheet in sheets:  # can process each sheet individually 
    df = pd.read_excel(base_in, sheet_name=sheet, header=None) # opens same file using pandas (to get correctly formatted dataframe)
    worksheet = workbook[sheet]
    rich_text_values = {}
    # Iterate over each cell in worksheet and extract font color
    for i, row in enumerate(worksheet.iter_rows()):
        for j, cell in enumerate(row):
            if cell.value is None or isinstance(cell.value, str) or isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            else:
                rich_text_values[i, j] = cell.value

    sequence_columns = set()
    for rkey, rval in rich_text_values.items():  # only update pandas cells if they require formatting, otherwise left as is
        sequence_columns.add(rkey[1])
        final_text = ''
        if isinstance(rval, int):  # not sure why integers sometimes get noted down as rich text, this line ignores them
            continue
        for text_block in rval:
            if isinstance(text_block, str):
                final_text += '\\seqsplit{%s}' % text_block  # seqsplit allow DNA sequence to be broken over multiple lines
            elif text_block.font.color is None or str(text_block.font.color.rgb) == "Values must be of type <class 'str'>":
                final_text += '\\seqsplit{%s}' % text_block.text
            else:
                final_text += '\\textcolor[HTML]{%s}{\seqsplit{%s}}' % (text_block.font.color.rgb[2:], text_block.text) # applies colour formatting
        if text_block.font.rFont == 'Courier':
            final_text = '\\texttt{%s}' % final_text  # texttt applies courier font
        df.iloc[rkey] = final_text

    for col in df.columns:  # manual check to ensure all Sequence columns get proper formatting applied
        if 'Sequence' in df[col].values:
            sequence_columns.add(col)

    # formatting check to catch any sequences that didn't have rich text (fully default)
    for col in sequence_columns:
        for ind, val in enumerate(df[col]):
            if not isinstance(val, float) and not isinstance(val, int) and 'seqsplit' not in val:
                df.iloc[ind, col] = '\\texttt{\\seqsplit{%s}}' % val

    df[df.isna().any(axis=1)] = ' ' # removes NaN values

    conv_text = df.to_latex(index=False, header=False).replace('_','\_')  # this is to prevent latex errors
    with open(base_out + sheet + '.txt', 'w') as f:
        f.write(conv_text)




# currently, I replace the tabular and headings this produces with the following (or similar):

# \begin{landscape}
#     \begin{longtable}{P{0.15\linewidth}p{0.53\linewidth}P{0.17\linewidth}P{0.1\linewidth}}
#         \caption[XXX]{XXX}
#         \label{table:clamp_variable}\\
#         \toprule
#         \textbf{Name} & \centering \textbf{Sequence} \arraybackslash & \textbf{Description} & \textbf{Length} \\
#         \midrule

#         \bottomrule
#     \end{longtable}
# \end{landscape}

# These packages are also needed in the latex document:

# \usepackage{seqsplit}
# \usepackage{longtable}
# \usepackage{tabularx}

